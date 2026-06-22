import csv
import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from common.models.analytics import ReportContent

# Adobe CID 字体，无需外部字体文件即可渲染简体中文。
_CJK_FONT = "STSong-Light"
_font_registered = False


def _ensure_font() -> None:
    global _font_registered
    if not _font_registered:
        pdfmetrics.registerFont(UnicodeCIDFont(_CJK_FONT))
        _font_registered = True


def render_csv(content: ReportContent) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([content.title])
    writer.writerow([])
    writer.writerow(["指标", "数值"])
    writer.writerows(content.summary)
    for section in content.sections:
        writer.writerow([])
        writer.writerow([section.title])
        writer.writerow(section.headers)
        writer.writerows(section.rows)
    # utf-8-sig 让 Excel 正确识别中文。
    return buffer.getvalue().encode("utf-8-sig")


def render_xlsx(content: ReportContent) -> bytes:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "概览"
    overview.append([content.title])
    overview.append([])
    overview.append(["指标", "数值"])
    for name, value in content.summary:
        overview.append([name, value])

    for index, section in enumerate(content.sections):
        # 工作表名上限 31 字符，超出会报错，做截断。
        sheet = workbook.create_sheet(title=section.title[:31] or f"section_{index}")
        sheet.append(section.headers)
        for row in section.rows:
            sheet.append(row)
        for col in range(1, len(section.headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def render_pdf(content: ReportContent) -> bytes:
    _ensure_font()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm)

    base = getSampleStyleSheet()["Normal"]
    title_style = ParagraphStyle("cjk_title", parent=base, fontName=_CJK_FONT, fontSize=16, spaceAfter=10)
    head_style = ParagraphStyle("cjk_head", parent=base, fontName=_CJK_FONT, fontSize=12, spaceBefore=8, spaceAfter=4)

    table_style = TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), _CJK_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3B52")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F4F8")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ])

    elements = [Paragraph(content.title, title_style)]
    elements.append(Paragraph("关键指标", head_style))
    elements.append(Table([["指标", "数值"]] + [list(item) for item in content.summary], style=table_style))

    for section in content.sections:
        elements.append(Paragraph(section.title, head_style))
        data = [section.headers] + (section.rows or [["（无数据）"] + [""] * (len(section.headers) - 1)])
        elements.append(Table(data, style=table_style))
        elements.append(Spacer(1, 4))

    doc.build(elements)
    return buffer.getvalue()


_RENDERERS = {"csv": render_csv, "xlsx": render_xlsx, "pdf": render_pdf}


def render(file_format: str, content: ReportContent) -> bytes:
    if file_format not in _RENDERERS:
        raise ValueError(f"不支持的导出格式：{file_format}")
    return _RENDERERS[file_format](content)
